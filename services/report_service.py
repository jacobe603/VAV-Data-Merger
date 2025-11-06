"""Service for generating Excel reports from TW2 data."""
import os
import logging
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font
from copy import copy

logger = logging.getLogger('vav_data_merger')


def generate_schedule_data_excel(tw2_data, project_name):
    """Generate Schedule Data Excel report from TW2 data using template.

    Args:
        tw2_data: List of dictionaries with TW2 data
        project_name: Project name to insert in the report

    Returns:
        BytesIO object containing the generated Excel file

    Raises:
        Exception: If template not found or Excel generation fails
    """
    try:
        # Load template file
        template_path = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                                     'templates', 'Schedule_Data_Template.xlsx')
        wb = load_workbook(template_path)
        ws = wb.active

        # IMPORTANT: Unmerge all cells that will be affected by row insertions and notes population
        # This must be done BEFORE we insert any rows
        merged_to_unmerge = []
        for merged_range in list(ws.merged_cells.ranges):
            # Unmerge anything in rows 6 and below (will be shifted by row insertions)
            # and anything in rows 14 and below (will be used for notes)
            if merged_range.min_row >= 6 or (merged_range.min_row >= 1 and merged_range.max_row >= 6):
                merged_to_unmerge.append(str(merged_range))

        for merged_range_str in merged_to_unmerge:
            try:
                ws.unmerge_cells(merged_range_str)
            except:
                pass

        # Update project name in row 2
        ws['A2'] = project_name

        # Get template row formatting (row 5)
        template_row = 5

        # Helper function to safely set cell value
        def safe_set_cell(cell_ref, value):
            ws[cell_ref].value = value

        # Insert data rows starting at row 5
        for i, record in enumerate(tw2_data):
            row_num = 5 + i

            # If not the first data row, insert a new row
            if i > 0:
                ws.insert_rows(row_num)

                # Copy formatting from template row to new row
                for col_num in range(1, 31):  # Columns A-AD
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(col_num)

                    source_cell = ws.cell(row=template_row, column=col_num)
                    target_cell = ws.cell(row=row_num, column=col_num)

                    # Copy formatting
                    if source_cell.font:
                        target_cell.font = copy(source_cell.font)
                    if source_cell.border:
                        target_cell.border = copy(source_cell.border)
                    if source_cell.alignment:
                        target_cell.alignment = copy(source_cell.alignment)
                    if source_cell.fill:
                        target_cell.fill = copy(source_cell.fill)
                    if source_cell.number_format:
                        target_cell.number_format = copy(source_cell.number_format)

            # Populate data columns
            try:
                safe_set_cell(f'A{row_num}', record.get('Tag', ''))
                safe_set_cell(f'F{row_num}', 'DESV')  # Model - always DESV
                safe_set_cell(f'G{row_num}', record.get('UnitSize', ''))
                safe_set_cell(f'H{row_num}', record.get('OutletSize', ''))
                safe_set_cell(f'I{row_num}', record.get('CFMDesign', ''))
                safe_set_cell(f'J{row_num}', record.get('CFMMinPrime', ''))
                safe_set_cell(f'K{row_num}', record.get('SPInlet', ''))
                safe_set_cell(f'L{row_num}', record.get('SPDownstream', ''))
                safe_set_cell(f'M{row_num}', record.get('SPMin', ''))
                safe_set_cell(f'N{row_num}', record.get('RadNCRoom', ''))
                safe_set_cell(f'O{row_num}', record.get('DisNCRoom', ''))
                safe_set_cell(f'P{row_num}', record.get('HWCFM', ''))

                if record.get('HWMBHCalc'):
                    safe_set_cell(f'Q{row_num}', round(float(record.get('HWMBHCalc', 0))))

                safe_set_cell(f'R{row_num}', record.get('HWEATCalc', ''))
                safe_set_cell(f'U{row_num}', record.get('HWEWT', ''))

                if record.get('HWLATCalc'):
                    safe_set_cell(f'V{row_num}', round(float(record.get('HWLATCalc', 0)), 1))

                if record.get('HWAPDCalc'):
                    safe_set_cell(f'W{row_num}', round(float(record.get('HWAPDCalc', 0)), 2))

                safe_set_cell(f'X{row_num}', record.get('HWGPMCalc', ''))

                if record.get('HWLWTCalc'):
                    safe_set_cell(f'Y{row_num}', round(float(record.get('HWLWTCalc', 0)), 1))

                if record.get('HWPDCalc'):
                    safe_set_cell(f'Z{row_num}', round(float(record.get('HWPDCalc', 0)), 2))

                hw_rows = record.get('HWRowsCalc') or record.get('HWRows', '')
                control_hand = record.get('ControlHand', '')
                if hw_rows:
                    safe_set_cell(f'AA{row_num}', f"{hw_rows}-{control_hand}")

                safe_set_cell(f'AB{row_num}', record.get('HWFPI', ''))
                safe_set_cell(f'AC{row_num}', record.get('ControlHand', ''))

                # Re-merge cells for this data row to match original template structure
                ws.merge_cells(f'A{row_num}:B{row_num}')
                ws.merge_cells(f'C{row_num}:E{row_num}')
                ws.merge_cells(f'R{row_num}:S{row_num}')

            except Exception as e:
                logger.error(f"Error processing row for tag {record.get('Tag', 'Unknown')}: {str(e)}")
                continue

        # Place notes section after data
        notes_start_row = 5 + len(tw2_data) + 2

        # Get fluid type info from first record
        fluid_type = tw2_data[0].get('FluidType', '') if tw2_data else ''
        pct_glycol = tw2_data[0].get('PctGlycol', 40) if tw2_data else 40

        if fluid_type == 'EG':
            fluid_description = f"{pct_glycol}% Ethylene Glycol"
        elif fluid_type == 'PG':
            fluid_description = f"{pct_glycol}% Propylene Glycol"
        elif fluid_type == 'Water' or fluid_type == '':
            fluid_description = "100% Water"
        else:
            fluid_description = f"{pct_glycol}% {fluid_type}"

        # Notes content
        notes = [
            ("Notes:", "1. Selections are based on Titus as Manufacturer."),
            (None, "2. All performance based on tests conducted in accordance with ASHRAE 130-2008 and AHRI 880-2011."),
            (None, "3. All NC levels determined using AHRI 885-2008 Appendix E."),
            (None, "4. All airflow, pressure losses and heating performance values have been corrected for altitude."),
            (None, "5. Units of measure: dimensions (in), airflow (cfm), water flow (gpm), air pressure (in wg), water head losses (ft) and temperatures (degF)."),
            (None, "6. Water pressure drop (WPd) units is in ft. water."),
            (None, f"7. Hot water performance based on {fluid_description}.")
        ]

        current_row = notes_start_row
        for label, note in notes:
            # Set row height to match other note rows
            ws.row_dimensions[current_row].height = 15

            if label:
                safe_set_cell(f'B{current_row}', label)
                ws[f'B{current_row}'].font = Font(name='Arial', size=8, bold=True)
                safe_set_cell(f'E{current_row}', note)
                # Merge notes label cells
                try:
                    ws.merge_cells(f'B{current_row}:C{current_row}')
                except:
                    pass
            else:
                safe_set_cell(f'E{current_row}', note)

            # Apply 8pt Arial font to note content
            ws[f'E{current_row}'].font = Font(name='Arial', size=8)

            # Merge notes content cells through column W for more room
            try:
                ws.merge_cells(f'E{current_row}:W{current_row}')
            except:
                pass
            current_row += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    except Exception as e:
        logger.exception(f"Error generating schedule data Excel: {str(e)}")
        raise
