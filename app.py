from flask import Flask, render_template, request, jsonify, Response, session, send_file
import logging
from flask_cors import CORS
from flask_session import Session
import pyodbc
import pandas as pd
import os
import json
import decimal
from datetime import datetime
from werkzeug.utils import secure_filename
import shutil
import tempfile
import time

app = Flask(__name__)
app.secret_key = 'vav-data-merger-secret-key-2025'  # Required for Flask sessions

# Configure Flask-Session for file-based session storage
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_FILE_DIR'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'sessions')
app.config['SESSION_PERMANENT'] = False
app.config['SESSION_USE_SIGNER'] = True
app.config['SESSION_KEY_PREFIX'] = 'vav-merger:'
Session(app)

# Basic logging configuration for app
logger = logging.getLogger('vav_data_merger')
if not logger.handlers:
    logger.setLevel(logging.INFO)
    _handler = logging.StreamHandler()
    _handler.setFormatter(logging.Formatter('%(asctime)s %(levelname)s %(message)s'))
logger.addHandler(_handler)


def _sanitize_path(p: str) -> str:
    """Strip surrounding quotes and whitespace from a filesystem path.
    Handles paths copied via Windows "Copy as path" which include quotes.
    """
    if not p:
        return p
    p = p.strip()
    if (p.startswith('"') and p.endswith('"')) or (p.startswith("'") and p.endswith("'")):
        return p[1:-1].strip()
    return p

CORS(app)

# Configure upload settings
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'tw2', 'mdb'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

# Ensure upload directory exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Session data is now handled by Flask sessions (removed global dict)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

class CustomJSONEncoder(json.JSONEncoder):
    """Custom JSON encoder to handle various data types"""
    def default(self, obj):
        if isinstance(obj, decimal.Decimal):
            return float(obj)
        elif isinstance(obj, datetime):
            return obj.isoformat()
        elif obj is None:
            return None
        elif isinstance(obj, (bytes, bytearray)):
            return str(obj, 'utf-8', errors='ignore')
        try:
            return super().default(obj)
        except TypeError:
            return str(obj)

def normalize_tag_format(tag):
    """Convert tag formats between Excel (V-1-1) and TW2 (V-1-01) formats"""
    if not tag or not isinstance(tag, str):
        return tag
    
    # Split the tag into parts
    parts = tag.split('-')
    if len(parts) >= 3:
        # Pad the last part with leading zero if needed
        try:
            last_num = int(parts[-1])
            if last_num < 10:
                parts[-1] = f"{last_num:02d}"
            return '-'.join(parts)
        except:
            return tag
    return tag

def safe_string_convert(value):
    """Safely convert any value to a JSON-safe string"""
    if value is None:
        return None
    elif pd.isna(value):  # Handle pandas NaN values
        return None
    elif isinstance(value, str):
        # Handle string NaN values too
        if str(value).lower() in ['nan', 'n/a', '']:
            return None
        # Remove any problematic characters
        return value.encode('ascii', 'ignore').decode('ascii')
    elif isinstance(value, (int, bool)):
        return value
    elif isinstance(value, float):
        # Handle float NaN values
        if pd.isna(value) or str(value).lower() == 'nan':
            return None
        return value
    elif isinstance(value, decimal.Decimal):
        return float(value)
    elif isinstance(value, datetime):
        return value.isoformat()
    else:
        # Convert to string and clean up
        try:
            str_val = str(value)
            if str_val.lower() in ['nan', 'n/a', '']:
                return None
            return str_val.encode('ascii', 'ignore').decode('ascii')
        except:
            return None

def get_mdb_connection(file_path):
    """Create a connection to the Access database with error handling"""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Database file not found: {file_path}")
    
    abs_path = os.path.abspath(file_path)
    
    # Try different connection strings
    connection_strings = [
        f'DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={abs_path};',
        f'DRIVER={{Microsoft Access Driver (*.mdb)}};DBQ={abs_path};',
        f'DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={abs_path};PWD=;'
    ]
    
    for conn_str in connection_strings:
        try:
            return pyodbc.connect(conn_str)
        except Exception as e:
            continue
    
    raise Exception(f"Failed to connect to database with any driver")

def read_tw2_data_safe(file_path):
    """Read TW2 data using safe methods that avoid cursor.columns()"""
    try:
        print(f"Attempting to read TW2 data from: {file_path}")
        
        conn = get_mdb_connection(file_path)
        cursor = conn.cursor()
        
        # Method 1: Get column names by running a dummy query
        print("Getting column information...")
        cursor.execute("SELECT * FROM tblSchedule WHERE 1=0")
        column_names = [desc[0] for desc in cursor.description]
        print(f"Found {len(column_names)} columns: {column_names[:10]}...")
        
        # Method 2: Get record count
        cursor.execute("SELECT COUNT(*) FROM tblSchedule")
        record_count = cursor.fetchone()[0]
        print(f"Found {record_count} records")
        
        # Method 3: Get actual data
        cursor.execute("SELECT * FROM tblSchedule")
        rows = cursor.fetchall()
        
        # Convert to safe dictionaries
        data = []
        for row in rows:
            row_dict = {}
            for i, column_name in enumerate(column_names):
                try:
                    # Safely convert each value
                    raw_value = row[i]
                    safe_value = safe_string_convert(raw_value)
                    row_dict[column_name] = safe_value
                except Exception as e:
                    print(f"Error converting column {column_name}: {e}")
                    row_dict[column_name] = None
            data.append(row_dict)
        
        conn.close()
        print("Successfully read TW2 data")
        
        return {
            'success': True,
            'data': data,
            'columns': column_names,
            'row_count': record_count
        }
        
    except Exception as e:
        print(f"Error reading TW2 data: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return {
            'success': False,
            'error': str(e).encode('ascii', 'ignore').decode('ascii')
        }

def clean_size_value(value):
    """Remove inch marks (") from size values and add zero-padding for numeric sizes"""
    if value is None or pd.isna(value):
        return value
    
    # Convert to string and remove all double quotes
    cleaned = str(value).replace('"', '')
    
    # Check if this is a simple numeric size that needs zero-padding
    # Handle both integer and string representations
    try:
        # Try to convert to integer to see if it's a simple number
        num_value = int(float(cleaned))
        # If it's a single digit (1-9), zero-pad to 2 digits
        if 1 <= num_value <= 9:
            return f"{num_value:02d}"
        # For larger numbers, return as string but ensure 2 digits minimum
        elif num_value >= 10:
            return str(num_value)
        else:
            return cleaned
    except (ValueError, TypeError):
        # Not a simple number - could be dimensions like "24x16" or "20x18"
        # Just return cleaned (without quotes)
        return cleaned

def normalize_header_text(text):
    """Clean and normalize header text"""
    if text is None or pd.isna(text):
        return ""
    # Convert to string and clean up
    cleaned = str(text).strip()
    # Remove line breaks and normalize whitespace - this is key for "UNIT\nNO."
    cleaned = ' '.join(cleaned.split())
    # Remove special characters that cause issues
    cleaned = cleaned.replace('&', 'and').replace('"', '').replace("'", "")
    return cleaned



def is_probably_header_value(value):
    """Heuristic to decide if a cell looks like part of a header row."""
    if value is None:
        return False

    if isinstance(value, (int, float)):
        return False

    text = str(value).strip()
    if not text:
        return False

    lower_text = text.lower()
    if lower_text in {'n/a', 'na', 'nan'}:
        return False

    if any(char.isdigit() for char in text):
        letters = sum(1 for c in text if c.isalpha())
        digits = sum(1 for c in text if c.isdigit())
        return letters > digits

    return True



def normalize_hw_rows_value(value):
    try:
        if value is None or value == '':
            return None
        if isinstance(value, (int, float)):
            return int(value)
        str_val = str(value).strip()
        if not str_val:
            return None
        return int(float(str_val))
    except Exception:
        return None

def combine_multi_row_headers(df, header_rows=2, title_row_offset=0):
    """Combine multi-row headers into single header row
    
    Args:
        df: DataFrame with raw Excel data
        header_rows: Number of header rows to combine (default 2)
        title_row_offset: Number of title rows to skip before headers (default 0)
    """
    headers = []

    start_row = title_row_offset
    end_row = title_row_offset + header_rows
    header_data = df.iloc[start_row:end_row]

    use_second_row = header_rows > 1 and len(header_data) > 1
    if use_second_row:
        second_row_values = [val for val in header_data.iloc[1].tolist() if not pd.isna(val)]
        if second_row_values:
            header_like_count = sum(1 for val in second_row_values if is_probably_header_value(val))
            if header_like_count < len(second_row_values) * 0.5:
                use_second_row = False

    for col_idx in range(len(df.columns)):
        row1_val = normalize_header_text(header_data.iloc[0, col_idx]) if not pd.isna(header_data.iloc[0, col_idx]) else ""
        row2_val = ""
        if use_second_row and not pd.isna(header_data.iloc[1, col_idx]):
            row2_val = normalize_header_text(header_data.iloc[1, col_idx])

        if row1_val and row2_val:
            combined = f"{row1_val}_{row2_val}"
        elif row1_val:
            combined = row1_val
        elif row2_val:
            combined = row2_val
        else:
            combined = f"Column_{col_idx + 1}"

        headers.append(combined)
    
    return headers

def map_excel_headers_to_standard(excel_headers):
    """Map Excel headers to our standard field names"""
    # Define mapping from combined Excel headers to standard names
    header_mapping = {
        # Unit identification
        'UNIT NO.': 'Unit_No',
        'UNIT_NO.': 'Unit_No', 
        'UNIT NO': 'Unit_No',
        'UNIT_NO': 'Unit_No',
        
        # Manufacturer
        'MANUFACTURER and MODEL NO.': 'Manufacturer_Model',
        'MANUFACTURER & MODEL NO.': 'Manufacturer_Model',
        'MANUFACTURER MODEL NO.': 'Manufacturer_Model',
        'MANUFACTURER and MODEL': 'Manufacturer_Model',
        
        # Unit size
        'UNIT SIZE': 'Unit_Size',
        'UNIT_SIZE': 'Unit_Size',
        
        # Dimensions
        'W x L x H': 'Dimensions',
        'Wx Lx H': 'Dimensions',
        'DIMENSIONS': 'Dimensions',
        
        # Inlet size
        'INLET SIZE': 'Inlet_Size',
        'INLET_SIZE': 'Inlet_Size',
        
        # Outlet size
        'OUTLET SIZE': 'Outlet_Size',
        'OUTLET_SIZE': 'Outlet_Size',
        
        # CFM values - handle the multi-row structure
        'CFM_MAX': 'CFM_Max',
        'CFM MAX': 'CFM_Max',
        'CFM_MIN': 'CFM_Min', 
        'CFM MIN': 'CFM_Min',
        'CFM_HEAT': 'CFM_Heat',
        'CFM HEAT': 'CFM_Heat',
        # Handle single CFM column that gets combined with sub-headers
        'CFM': 'CFM_Max',  # Default CFM to Max if no sub-header
        
        # Temperature values
        'EAT': 'EAT',
        'LAT': 'LAT',
        
        # Other values
        'MBH': 'MBH',
        'TOTAL MBH': 'Total_MBH',
        'TOTAL_MBH': 'Total_MBH',
        'EWT': 'EWT',
        'FLUID': 'Fluid',
        'GPM': 'GPM',
        'MAX WPD': 'Max_WPD',
        'MAX_WPD': 'Max_WPD',
        'WPD': 'WPD',
        'APD': 'APD',
        'NOTES': 'Notes',
        'TAG': 'Unit_No'
    }
    
    # Map headers with smart CFM handling
    mapped_headers = []
    
    for i, header in enumerate(excel_headers):
        header_upper = header.upper().strip()
        
        # Direct mapping first
        if header_upper in header_mapping:
            mapped_headers.append(header_mapping[header_upper])
        # Handle special cases for CFM sub-columns
        elif header_upper == 'MAX':
            # Check if this is likely a CFM sub-column by looking at previous headers
            if i > 0 and ('CFM' in excel_headers[i-1].upper() or any('CFM' in str(excel_headers[j]).upper() for j in range(max(0, i-3), i))):
                mapped_headers.append('CFM_Max')
            else:
                mapped_headers.append('MAX')
        elif header_upper == 'MIN':
            # Check if this is likely a CFM sub-column
            if i > 0 and ('CFM' in excel_headers[i-1].upper() or any('CFM' in str(excel_headers[j]).upper() for j in range(max(0, i-3), i))):
                mapped_headers.append('CFM_Min')
            else:
                mapped_headers.append('MIN')
        elif header_upper == 'HEAT':
            # Check if this is likely a CFM sub-column
            if i > 0 and ('CFM' in excel_headers[i-1].upper() or any('CFM' in str(excel_headers[j]).upper() for j in range(max(0, i-3), i))):
                mapped_headers.append('CFM_Heat')
            else:
                mapped_headers.append('HEAT')
        # Handle empty or generic CFM columns
        elif header_upper == 'CFM' or (header_upper == '' and i > 0 and 'CFM' in excel_headers[i-1].upper()):
            # This is likely the start of CFM multi-column, assign based on position
            mapped_headers.append('CFM_Max')  # First CFM column is usually MAX
        # Default handling - keep original or create generic name
        else:
            # Try partial matching for common patterns
            if 'MANUFACTURER' in header_upper:
                mapped_headers.append('Manufacturer_Model')
            elif 'UNIT' in header_upper and ('SIZE' in header_upper or 'NO' in header_upper):
                if 'SIZE' in header_upper:
                    mapped_headers.append('Unit_Size')
                else:
                    mapped_headers.append('Unit_No')
            elif 'INLET' in header_upper:
                mapped_headers.append('Inlet_Size')
            elif 'OUTLET' in header_upper:
                mapped_headers.append('Outlet_Size')
            elif 'DIMENSION' in header_upper or 'x' in header_upper.lower():
                mapped_headers.append('Dimensions')
            else:
                # Keep original header but clean it up
                clean_header = header.replace(' ', '_').replace('&', 'and')
                mapped_headers.append(clean_header)
    
    print(f"Header mapping result: {list(zip(excel_headers, mapped_headers))}")
    return mapped_headers

def read_excel_data_safe(file_path, data_start_row=3, header_rows=2, skip_title_row=True):
    """Read Excel data with proper error handling and configurable header detection
    
    Args:
        file_path: Path to Excel file
        data_start_row: Row number where data starts (1-based)
        header_rows: Number of header rows to combine
        skip_title_row: Whether to skip the first row as title
    """
    try:
        print("=" * 50)
        print("EXCEL FILE PROCESSING STARTED")
        print("=" * 50)
        print(f"Attempting to read Excel data from: {file_path}")
        
        # Read the Excel file without any header assumptions
        df_raw = pd.read_excel(file_path, sheet_name=0, header=None)
        print(f"Raw Excel shape: {df_raw.shape}")
        
        # Show first 5 rows for debugging
        print("=== FIRST 5 ROWS OF RAW EXCEL ===")
        for i in range(min(5, len(df_raw))):
            row_values = df_raw.iloc[i].tolist()
            print(f"Row {i}: {row_values}")
        print("=== END RAW EXCEL PREVIEW ===")
        
        # Use configurable header detection
        title_row_offset = 1 if skip_title_row else 0
        
        print(f"Configuration - Data start row: {data_start_row}, Header rows: {header_rows}, Skip title: {skip_title_row}")
        print(f"Title row offset: {title_row_offset}")
        
        # Auto-adjust title row offset if the first row is actual headers
        first_row_values = [val for val in df_raw.iloc[0].tolist() if not pd.isna(val)]
        if skip_title_row and first_row_values:
            header_like = sum(1 for val in first_row_values if is_probably_header_value(val))
            if header_like >= max(1, len(first_row_values) // 2):
                title_row_offset = 0
                print('AUTO-DETECT: using first row as headers')

        # Combine multi-row headers based on configuration
        excel_headers = combine_multi_row_headers(df_raw, header_rows=header_rows, title_row_offset=title_row_offset)
        print(f"Combined headers detected: {excel_headers}")
        
        # Map Excel headers to our standard names
        mapped_headers = map_excel_headers_to_standard(excel_headers)
        print(f"Mapped to standard headers: {mapped_headers}")
        
        # Extract data starting from the configured row (convert to 0-based index)
        data_start_index = data_start_row - 1
        print(f"Extracting data starting from row {data_start_row} (index {data_start_index})")
        df_data = df_raw.iloc[data_start_index:].reset_index(drop=True)
        df_data.columns = mapped_headers[:len(df_data.columns)]
        
        # Remove completely empty rows
        df_cleaned = df_data.dropna(how='all').reset_index(drop=True)
        
        # Headers should already be properly mapped, so Unit_No column should exist
        # Don't try to detect or recreate Unit_No column - trust the header mapping
        print(f"DataFrame columns after mapping: {list(df_cleaned.columns)}")
        if 'Unit_No' in df_cleaned.columns:
            print(f"Unit_No column found with sample values: {df_cleaned['Unit_No'].head().tolist()}")
        else:
            print("WARNING: Unit_No column not found in mapped headers")
        
        # Remove rows where critical data is missing (check multiple columns)
        # Keep rows that have data in at least Unit_Size or CFM_Max or Manufacturer_Model
        critical_cols = ['Unit_Size', 'CFM_Max', 'Manufacturer_Model']
        available_critical = [col for col in critical_cols if col in df_cleaned.columns]
        
        if available_critical:
            # Keep rows that have data in at least one critical column
            mask = df_cleaned[available_critical].notna().any(axis=1)
            df_cleaned = df_cleaned[mask].reset_index(drop=True)
        
        # Clean size fields - remove inch marks
        size_columns = ['Unit_Size', 'Inlet_Size', 'Outlet_Size']
        for col in size_columns:
            if col in df_cleaned.columns:
                df_cleaned[col] = df_cleaned[col].apply(clean_size_value)
        
        # Convert to safe format
        data = []
        for _, row in df_cleaned.iterrows():
            row_dict = {}
            for col in df_cleaned.columns:
                raw_value = row[col]
                safe_value = safe_string_convert(raw_value)
                row_dict[col] = safe_value
            data.append(row_dict)
        
        print(f"Successfully read {len(data)} Excel records")
        
        return {
            'success': True,
            'data': data,
            'columns': list(df_cleaned.columns),
            'row_count': len(data),
            'header_info': {
                'original_headers': excel_headers,
                'combined_headers': excel_headers,  # Same as original for now
                'mapped_headers': mapped_headers,
                'data_start_row': data_start_row
            }
        }
        
    except Exception as e:
        print(f"Error reading Excel data: {str(e)}")
        return {
            'success': False,
            'error': str(e).encode('ascii', 'ignore').decode('ascii')
        }


def reload_tw2_data_from_disk(preferred_paths=None):
    """Reload TW2 data from disk, updating the session with the latest contents."""
    candidates = []
    seen_paths = set()

    def _add_candidate(label, candidate_path):
        if not candidate_path:
            return
        sanitized = _sanitize_path(str(candidate_path).strip())
        if not sanitized:
            return
        normalized = os.path.abspath(sanitized)
        if normalized in seen_paths:
            return
        candidates.append((label, sanitized))
        seen_paths.add(normalized)

    if preferred_paths:
        for label, candidate_path in preferred_paths:
            _add_candidate(label, candidate_path)

    _add_candidate('original', session.get('original_tw2_path'))
    _add_candidate('local', session.get('updated_tw2_path'))
    _add_candidate('tw2', session.get('tw2_file'))

    if not candidates:
        return {'success': False, 'error': 'No TW2 file path available', 'code': 404}

    last_error = {'message': 'Unable to reload TW2 data', 'code': 500}

    for label, candidate_path in candidates:
        if not os.path.exists(candidate_path):
            last_error = {'message': f'File not found at {candidate_path}', 'code': 404}
            continue

        result = read_tw2_data_safe(candidate_path)
        if result.get('success'):
            session['updated_tw2_data'] = result['data']
            session['updated_tw2_columns'] = result['columns']
            session['updated_tw2_records'] = result['row_count']
            session['updated_tw2_filename'] = os.path.basename(candidate_path)
            session['last_tw2_reload_path'] = candidate_path
            session['last_tw2_reload_source'] = label
            session['tw2_last_path'] = candidate_path
            try:
                session['tw2_last_mtime'] = os.path.getmtime(candidate_path)
            except Exception:
                session.pop('tw2_last_mtime', None)
            return {
                'success': True,
                'path': candidate_path,
                'source': label,
                'row_count': result['row_count'],
                'column_count': len(result['columns'])
            }
        else:
            last_error = {
                'message': result.get('error', 'Unknown error while reading TW2 file'),
                'code': 500
            }

    return {'success': False, 'error': last_error['message'], 'code': last_error['code']}



def normalize_unit_tag(tag):
    """Normalize unit tags by adding zero-padding: V-1-1 -> V-1-01, V-1-12 -> V-1-12"""
    if not tag:
        return tag
    
    tag_str = str(tag).strip()
    
    # Look for pattern like V-1-1, V-2-3, etc.
    import re
    pattern = r'^([A-Z]+-\d+-)(\d+)$'
    match = re.match(pattern, tag_str)
    
    if match:
        prefix = match.group(1)  # V-1-
        number = match.group(2)  # 1
        
        # Pad single digits with zero
        if len(number) == 1:
            return f"{prefix}{number.zfill(2)}"
    
    return tag_str

def compare_performance_data(excel_data, updated_tw2_data, mbh_lat_lower_margin=15, mbh_lat_upper_margin=25, wpd_threshold=5, apd_threshold=0.25):
    """Compare performance values between Excel and updated TW2 data"""
    try:
        comparison_results = []
        
        # Convert updated TW2 data to DataFrame for easier processing
        updated_tw2_df = pd.DataFrame(updated_tw2_data)
        excel_df = pd.DataFrame(excel_data)
        
        # Create index mapping for faster lookups with normalized tags
        tw2_index = {}
        for row in updated_tw2_data:
            original_tag = str(row.get('Tag', '')).strip()
            normalized_tag = normalize_unit_tag(original_tag)
            tw2_index[normalized_tag] = row
            # Also keep original tag as backup
            tw2_index[original_tag] = row
        
        for excel_row in excel_data:
            unit_tag = str(excel_row.get('Unit_No', '')).strip()
            if not unit_tag:
                continue
                
            # Normalize the Excel unit tag for matching
            normalized_excel_tag = normalize_unit_tag(unit_tag)
            
            # Find matching TW2 record - try normalized first, then original
            tw2_row = tw2_index.get(normalized_excel_tag) or tw2_index.get(unit_tag)
            if not tw2_row:
                comparison_results.append({
                    'unit_tag': f"{unit_tag} \u001a {normalized_excel_tag}" if normalized_excel_tag != unit_tag else unit_tag,
                    'status': 'Not Found',
                    'excel_mbh': excel_row.get('MBH', 'N/A'),
                    'tw2_mbh': 'N/A',
                    'mbh_diff': 'N/A',
                    'excel_lat': excel_row.get('LAT', 'N/A'),
                    'tw2_lat': 'N/A',
                    'lat_diff': 'N/A',
                    'tw2_wpd': 'N/A',
                    'tw2_apd': 'N/A',
                    'tw2_hw_rows': None,
                    'tw2_hw_rows_raw': None,
                })
                continue
            
            # Extract values
            excel_mbh = excel_row.get('MBH') or excel_row.get('MBH_Total') or excel_row.get('Total_MBH')
            excel_lat = excel_row.get('LAT') or excel_row.get('Leaving_Air_Temp')
            
            tw2_mbh = tw2_row.get('HWMBHCalc')
            tw2_lat = tw2_row.get('HWLATCalc')
            tw2_wpd = tw2_row.get('HWPDCalc')
            tw2_apd = tw2_row.get('HWAPDCalc')

            tw2_hw_raw = None
            for hw_key in ('HWRowsCalc', 'HWRows', 'HWRow'):
                candidate = tw2_row.get(hw_key)
                if candidate not in (None, ''):
                    tw2_hw_raw = candidate
                    break
            tw2_hw_rows = normalize_hw_rows_value(tw2_hw_raw)

            
            # Calculate differences and status
            mbh_diff = 'N/A'
            lat_diff = 'N/A'
            status_flags = []
            
            # MBH comparison with separate upper/lower margins
            if excel_mbh is not None and tw2_mbh is not None:
                try:
                    excel_mbh_val = float(excel_mbh)
                    tw2_mbh_val = float(tw2_mbh)
                    if excel_mbh_val != 0:
                        mbh_diff = ((tw2_mbh_val - excel_mbh_val) / excel_mbh_val) * 100
                        # Check if outside acceptable range: -15% to +25%
                        if mbh_diff < -mbh_lat_lower_margin:  # Too low (under by more than 15%)
                            status_flags.append(f'MBH {mbh_diff:.1f}% (too low)')
                        elif mbh_diff > mbh_lat_upper_margin:  # Too high (over by more than 25%)
                            status_flags.append(f'MBH {mbh_diff:.1f}% (too high)')
                except (ValueError, TypeError):
                    pass
            
            # LAT comparison with separate upper/lower margins
            if excel_lat is not None and tw2_lat is not None:
                try:
                    excel_lat_val = float(excel_lat)
                    tw2_lat_val = float(tw2_lat)
                    if excel_lat_val != 0:
                        lat_diff = ((tw2_lat_val - excel_lat_val) / excel_lat_val) * 100
                        # Check if outside acceptable range: -15% to +25%
                        if lat_diff < -mbh_lat_lower_margin:  # Too low (under by more than 15%)
                            status_flags.append(f'LAT {lat_diff:.1f}% (too low)')
                        elif lat_diff > mbh_lat_upper_margin:  # Too high (over by more than 25%)
                            status_flags.append(f'LAT {lat_diff:.1f}% (too high)')
                except (ValueError, TypeError):
                    pass
            
            # WPD check
            if tw2_wpd is not None:
                try:
                    wpd_val = float(tw2_wpd)
                    if wpd_val > wpd_threshold:
                        status_flags.append(f'WPD {wpd_val:.2f}')
                except (ValueError, TypeError):
                    pass
            
            # APD check
            if tw2_apd is not None:
                try:
                    apd_val = float(tw2_apd)
                    if apd_val > apd_threshold:
                        status_flags.append(f'APD {apd_val:.2f}')
                except (ValueError, TypeError):
                    pass
            
            # Determine overall status
            if status_flags:
                status = 'Fail' if any('MBH' in flag or 'LAT' in flag for flag in status_flags) else 'Warning'
            else:
                status = 'Pass'
            
            comparison_results.append({
                'unit_tag': f"{unit_tag} \u001a {normalized_excel_tag}" if normalized_excel_tag != unit_tag else unit_tag,
                'status': status,
                'status_details': ', '.join(status_flags) if status_flags else 'All within range',
                'excel_mbh': excel_mbh,
                'tw2_mbh': tw2_mbh,
                'mbh_diff': f'{mbh_diff:.1f}%' if isinstance(mbh_diff, (int, float)) else mbh_diff,
                'excel_lat': excel_lat,
                'tw2_lat': tw2_lat,
                'lat_diff': f'{lat_diff:.1f}%' if isinstance(lat_diff, (int, float)) else lat_diff,
                'tw2_wpd': tw2_wpd,
                'tw2_apd': tw2_apd,
                'tw2_hw_rows': tw2_hw_rows,
                'tw2_hw_rows_raw': tw2_hw_raw if tw2_hw_raw not in (None, '') else None
            })
        
        return {
            'success': True,
            'results': comparison_results,
            'summary': {
                'total': len(comparison_results),
                'pass': len([r for r in comparison_results if r['status'] == 'Pass']),
                'warning': len([r for r in comparison_results if r['status'] == 'Warning']),
                'fail': len([r for r in comparison_results if r['status'] == 'Fail']),
                'not_found': len([r for r in comparison_results if r['status'] == 'Not Found'])
            }
        }
    
    except Exception as e:
        return {
            'success': False,
            'error': f'Error during performance comparison: {str(e)}'
        }

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload_tw2', methods=['POST'])
def upload_tw2():
    """Upload and analyze .tw2 file with fixed encoding handling"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            
            # Save the file
            file.save(filepath)
            abs_filepath = os.path.abspath(filepath)
            
            # Read the tw2 data using safe method
            result = read_tw2_data_safe(abs_filepath)
            
            if result['success']:
                # Store in session data
                session['tw2_file'] = abs_filepath
                session['tw2_data'] = result['data']
                session['tw2_columns'] = result['columns']
                session['original_filename'] = file.filename
            
            # Use custom JSON encoding
            return Response(
                json.dumps(result, cls=CustomJSONEncoder, ensure_ascii=True),
                mimetype='application/json'
            )
            
    except Exception as e:
        error_msg = str(e).encode('ascii', 'ignore').decode('ascii')
        return Response(
            json.dumps({'success': False, 'error': error_msg}, ensure_ascii=True),
            mimetype='application/json',
            status=500
        )

@app.route('/upload_excel', methods=['POST'])
def upload_excel():
    """Upload and analyze Excel file"""
    try:
        print("=" * 50)
        print("UPLOAD_EXCEL ROUTE CALLED")
        print("=" * 50)
        
        if 'file' not in request.files:
            print("ERROR: No file in request")
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            # Get configuration parameters from form data
            data_start_row = int(request.form.get('data_start_row', 3))
            header_rows = int(request.form.get('header_rows', 2))
            skip_title_row = request.form.get('skip_title_row', 'true').lower() == 'true'
            
            # Read the Excel data with configuration
            result = read_excel_data_safe(filepath, data_start_row=data_start_row, 
                                        header_rows=header_rows, skip_title_row=skip_title_row)
            
            if result['success']:
                # Store in session data
                session['excel_file'] = filepath
                session['excel_data'] = result['data']
                session['excel_columns'] = result['columns']
            
            return Response(
                json.dumps(result, cls=CustomJSONEncoder, ensure_ascii=True),
                mimetype='application/json'
            )
            
    except Exception as e:
        error_msg = str(e).encode('ascii', 'ignore').decode('ascii')
        return Response(
            json.dumps({'success': False, 'error': error_msg}, ensure_ascii=True),
            mimetype='application/json',
            status=500
        )

@app.route('/debug_excel', methods=['GET'])
def debug_excel():
    """Debug endpoint to show raw Excel data"""
    if 'excel_file' not in session:
        return jsonify({'error': 'No Excel file uploaded'}), 400
    
    try:
        file_path = session['excel_file']
        df_raw = pd.read_excel(file_path, sheet_name=0, header=None)
        
        # Get first 5 rows as lists
        debug_info = {
            'file_path': file_path,
            'shape': list(df_raw.shape),
            'first_5_rows': []
        }
        
        for i in range(min(5, len(df_raw))):
            row_data = df_raw.iloc[i].fillna('').tolist()
            debug_info['first_5_rows'].append({
                'row_index': i,
                'data': row_data
            })
        
        return jsonify(debug_info)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/debug_headers', methods=['GET'])
def debug_headers():
    """Debug endpoint to show header processing"""
    if 'excel_file' not in session:
        return jsonify({'error': 'No Excel file uploaded'}), 400
    
    try:
        file_path = session['excel_file'] 
        df_raw = pd.read_excel(file_path, sheet_name=0, header=None)
        
        # Show header processing step by step
        title_row_offset = 1  # Skip title row
        header_rows = 2
        
        # Get raw header rows
        start_row = title_row_offset 
        end_row = title_row_offset + header_rows
        header_data = df_raw.iloc[start_row:end_row]
        
        debug_info = {
            'title_row_offset': title_row_offset,
            'header_rows': header_rows,
            'raw_header_row_1': header_data.iloc[0].fillna('').tolist(),
            'raw_header_row_2': header_data.iloc[1].fillna('').tolist(),
        }
        
        # Process headers
        excel_headers = combine_multi_row_headers(df_raw, header_rows=header_rows, title_row_offset=title_row_offset)
        mapped_headers = map_excel_headers_to_standard(excel_headers)
        
        debug_info['combined_headers'] = excel_headers
        debug_info['mapped_headers'] = mapped_headers
        debug_info['header_mapping'] = list(zip(excel_headers, mapped_headers))
        
        return jsonify(debug_info)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/debug_data', methods=['GET'])
def debug_data():
    """Debug endpoint to show data extraction"""
    if 'excel_file' not in session:
        return jsonify({'error': 'No Excel file uploaded'}), 400
    
    try:
        file_path = session['excel_file'] 
        df_raw = pd.read_excel(file_path, sheet_name=0, header=None)
        
        # Simulate the data extraction with default settings
        data_start_row = 4  # This should be row 4 (index 3)
        data_start_index = data_start_row - 1  # Convert to 0-based = 3
        
        debug_info = {
            'data_start_row': data_start_row,
            'data_start_index': data_start_index,
            'total_rows': len(df_raw),
            'extracted_data_shape': None,
            'first_3_extracted_rows': []
        }
        
        # Extract data
        df_data = df_raw.iloc[data_start_index:].reset_index(drop=True)
        debug_info['extracted_data_shape'] = list(df_data.shape)
        
        # Show first 3 rows of extracted data
        for i in range(min(3, len(df_data))):
            row_data = df_data.iloc[i].fillna('').tolist()
            debug_info['first_3_extracted_rows'].append({
                'extracted_row_index': i,
                'original_row_index': data_start_index + i,
                'data': row_data
            })
        
        return jsonify(debug_info)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/get_mapping_fields', methods=['GET'])
def get_mapping_fields():
    """Get the fields available for mapping from both files"""
    
    # Target fields in tblSchedule that need to be mapped
    target_fields = [
        'Tag', 'UnitSize', 'InletSize', 'CFMDesign', 'CFMMinPrime',
        'HWCFM', 'HWGPM', 'HeatingPrimaryAirflow', 'CFMMin'
    ]
    
    # Get tw2 columns if available
    tw2_fields = []
    if 'tw2_columns' in session:
        tw2_fields = session['tw2_columns']
    
    # Get Excel columns if available
    excel_fields = []
    if 'excel_columns' in session:
        excel_fields = session['excel_columns']
    
    # Suggested mappings
    suggested_mappings = {
        'Tag': 'Unit_No',
        'UnitSize': 'Unit_Size',
        'InletSize': 'Unit_Size',  # Both UnitSize and InletSize map to Unit_Size with special logic
        'CFMDesign': 'CFM_Max',
        'CFMMinPrime': 'CFM_Min',
        'CFMMin': 'CFM_Min',  # Alternative field name for CFM Min
        'HeatingPrimaryAirflow': 'CFM_Heat',  # Alternative field name for heating airflow
        'HWCFM': 'CFM_Heat',
        'HWGPM': 'GPM'
    }
    
    result = {
        'target_fields': target_fields,
        'tw2_fields': tw2_fields,
        'excel_fields': excel_fields,
        'suggested_mappings': suggested_mappings
    }
    
    return Response(
        json.dumps(result, cls=CustomJSONEncoder, ensure_ascii=True),
        mimetype='application/json'
    )

@app.route('/upload_updated_tw2', methods=['POST'])
def upload_updated_tw2():
    """Handle updated TW2 file upload for performance comparison"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
            
        if not file.filename.lower().endswith(('.tw2', '.mdb')):
            return jsonify({'error': 'Please upload a TW2 or MDB file'}), 400
        
        # Get optional original file path from form data
        original_path = _sanitize_path(request.form.get('original_path', '').strip())
        print(f"UPLOAD: Original path provided: {original_path}")
        
        # Save file persistently for refresh functionality
        filename = secure_filename(file.filename)
        
        # Create a persistent uploads directory
        upload_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
        os.makedirs(upload_dir, exist_ok=True)
        
        # Save with timestamp to avoid conflicts
        timestamp = str(int(time.time()))
        persistent_filename = f"{timestamp}_{filename}"
        persistent_path = os.path.join(upload_dir, persistent_filename)
        file.save(persistent_path)
        
        # Read the updated TW2 data
        result = read_tw2_data_safe(persistent_path)
        
        if result['success']:
            # Store updated TW2 data and file path in session
            session['updated_tw2_data'] = result['data']
            session['updated_tw2_columns'] = result['columns']
            session['updated_tw2_filename'] = filename
            session['updated_tw2_records'] = result['row_count']
            session['updated_tw2_path'] = persistent_path  # Local copy for refresh
            
            # Store original path if provided for remote refresh capability
            if original_path:
                session['original_tw2_path'] = original_path
                print(f"UPLOAD: Stored original path in session: {original_path}")
            else:
                # Clear original path if not provided
                session.pop('original_tw2_path', None)
                print("UPLOAD: No original path provided, cleared from session")
            
            return jsonify({
                'success': True,
                'filename': filename,
                'records': result['row_count'],
                'columns': result['columns'][:10],  # First 10 columns for display
                'column_count': len(result['columns']),
                'message': f'Successfully read {result["row_count"]} records with {len(result["columns"])} columns'
            })
        else:
            return jsonify({'error': f'Failed to read TW2 file: {result["error"]}'}), 400
            
    except Exception as e:
        print(f"Error in upload_updated_tw2: {str(e)}")
        return jsonify({'error': f'Error processing updated TW2 file: {str(e)}'}), 500

@app.route('/apply_mapping', methods=['POST'])
def apply_mapping():
    """Apply the mapping and update the tw2 database"""
    try:
        data = request.json
        mappings = data.get('mappings', {})
        
        if not session.get('tw2_file') or not session.get('excel_data'):
            return Response(
                json.dumps({'success': False, 'error': 'Files not loaded'}, ensure_ascii=True),
                mimetype='application/json',
                status=400
            )
        
        # Create a backup
        backup_path = session['tw2_file'] + '.backup_' + datetime.now().strftime('%Y%m%d_%H%M%S')
        shutil.copy2(session['tw2_file'], backup_path)
        
        # Connect to the database
        conn = get_mdb_connection(session['tw2_file'])
        cursor = conn.cursor()
        
        updated_records = 0
        errors = []
        
        # Process each Excel row
        for excel_row in session['excel_data']:
            try:
                # Get the Tag value from Excel to match with tw2
                tag_field = mappings.get('Tag')
                if not tag_field:
                    continue
                    
                tag_value = excel_row.get(tag_field)
                if not tag_value:
                    continue
                
                # Normalize tag format for matching (V-1-1 -> V-1-01)
                normalized_tag = normalize_tag_format(str(tag_value))
                print(f"Original tag: {tag_value} -> Normalized: {normalized_tag}")
                
                # Implement batched field updates to avoid SQL parameter limits
                print(f"Debug - All mappings received: {mappings}")
                
                # Group fields into smaller batches to isolate the problematic field
                field_batches = [
                    ['UnitSize', 'InletSize', 'CFMDesign'],      # Batch 1: Size and design fields
                    ['CFMMinPrime', 'CFMMin'],                   # Batch 2a: CFM Min fields
                    ['HWCFM', 'HeatingPrimaryAirflow'],                 # Batch 2b: Heating airflow fields
                    ['HWGPM']                                    # Batch 3: GPM field
                ]
                
                record_updated = False
                batch_success_count = 0
                
                for batch_num, batch_fields in enumerate(field_batches, 1):
                    update_fields = []
                    params = []
                    
                    # Build update for this batch
                    for tw2_field in batch_fields:
                        if tw2_field in mappings:
                            excel_field = mappings[tw2_field]
                            if excel_field in excel_row:
                                value = excel_row[excel_field]
                                
                                # Special handling for Unit_Size mapping
                                if excel_field == 'Unit_Size' and tw2_field in ['UnitSize', 'InletSize']:
                                    # Clean and format the value first
                                    cleaned_value = clean_size_value(value)
                                    
                                    if tw2_field == 'UnitSize':
                                        # UnitSize always gets the cleaned Unit_Size value
                                        final_value = cleaned_value
                                    elif tw2_field == 'InletSize':
                                        # InletSize: special case when Unit_Size = 40, then InletSize = "24x16"
                                        # Otherwise, InletSize gets the same value as UnitSize
                                        try:
                                            unit_size_num = int(float(str(cleaned_value)))
                                            if unit_size_num == 40:
                                                final_value = "24x16"
                                            else:
                                                final_value = cleaned_value
                                        except (ValueError, TypeError):
                                            # If not a number, use cleaned value as-is
                                            final_value = cleaned_value
                                    
                                    update_fields.append(f"[{tw2_field}] = ?")
                                    if final_value is None or (isinstance(final_value, str) and str(final_value).strip() == ''):
                                        params.append(None)
                                    else:
                                        params.append(final_value)
                                    print(f"Debug - Batch {batch_num}: Unit_Size special mapping {tw2_field} = {final_value} (from {value})")
                                    
                                else:
                                    # Standard field mapping - apply size cleaning if it's a size field
                                    if tw2_field in ['UnitSize', 'InletSize'] or 'Size' in tw2_field:
                                        cleaned_value = clean_size_value(value)
                                        final_value = cleaned_value
                                    else:
                                        final_value = value
                                    
                                    update_fields.append(f"[{tw2_field}] = ?")
                                    # Convert empty strings to None for database
                                    if final_value is None or (isinstance(final_value, str) and str(final_value).strip() == ''):
                                        params.append(None)
                                    else:
                                        params.append(final_value)
                                    print(f"Debug - Batch {batch_num}: Adding field {tw2_field} = {final_value}")
                    
                    # Execute batch if we have fields to update
                    if update_fields:
                        # Add the WHERE clause parameter - use normalized tag for matching
                        params.append(normalized_tag)
                        
                        query = f"""
                            UPDATE tblSchedule 
                            SET {', '.join(update_fields)}
                            WHERE [Tag] = ?
                        """
                        
                        print(f"Debug - Batch {batch_num} query: {query}")
                        print(f"Debug - Batch {batch_num} params: {params}")
                        
                        try:
                            cursor.execute(query, params)
                            if cursor.rowcount > 0:
                                batch_success_count += 1
                                record_updated = True
                                print(f"Debug - Batch {batch_num} successful for {tag_value}")
                            else:
                                print(f"Debug - Batch {batch_num} no rows affected for {tag_value}")
                        except Exception as batch_error:
                            error_msg = f"Batch {batch_num} error for {tag_value}: {str(batch_error)}"
                            errors.append(error_msg)
                            print(error_msg)
                
                if record_updated:
                    updated_records += 1
                    print(f"Debug - Successfully updated {tag_value} with {batch_success_count}/{len(field_batches)} batches")
                    
            except Exception as e:
                error_msg = f"Error updating {tag_value}: {str(e)}"
                errors.append(error_msg)
                print(error_msg)
        
        # Commit the changes
        conn.commit()
        conn.close()
        
        result = {
            'success': True,
            'updated_records': updated_records,
            'backup_file': backup_path,
            'errors': errors if errors else None
        }
        
        return Response(
            json.dumps(result, cls=CustomJSONEncoder, ensure_ascii=True),
            mimetype='application/json'
        )
        
    except Exception as e:
        error_msg = str(e).encode('ascii', 'ignore').decode('ascii')
        return Response(
            json.dumps({'success': False, 'error': error_msg}, ensure_ascii=True),
            mimetype='application/json',
            status=500
        )


@app.route('/download_merged_tw2', methods=['GET'])
def download_merged_tw2():
    """Download the merged TW2 file"""
    try:
        if not session.get('tw2_file'):
            return jsonify({'error': 'No TW2 file available for download'}), 400

        tw2_file_path = session['tw2_file']
        if not os.path.exists(tw2_file_path):
            return jsonify({'error': 'TW2 file not found'}), 404

        original_filename = os.path.basename(tw2_file_path)
        name_part, ext = os.path.splitext(original_filename)
        download_name = f"{name_part}_merged{ext}"

        return send_file(
            tw2_file_path,
            as_attachment=True,
            download_name=download_name,
            mimetype='application/octet-stream'
        )
    except Exception as e:
        return jsonify({'error': f'Failed to download TW2 file: {str(e)}'}), 500


@app.route('/save_hw_rows', methods=['POST'])
def save_hw_rows():
    """Save HW Rows changes to the TW2 file"""
    try:
        data = request.get_json(silent=True) or {}
        edits = data.get('edits', [])

        incoming_path = data.get('original_path')
        if incoming_path is not None:
            sanitized_path = _sanitize_path(str(incoming_path).strip())
            if sanitized_path:
                session['original_tw2_path'] = sanitized_path
            else:
                return jsonify({
                    'success': False,
                    'error': 'Original TW2 file path is required before saving HW Rows.'
                }), 400

        if not edits:
            return jsonify({'success': False, 'error': 'No edits provided'}), 400

        for edit in edits:
            hw_rows = edit.get('hw_rows')
            if hw_rows not in [1, 2, 3, 4]:
                return jsonify({'success': False, 'error': f'Invalid HW Rows value: {hw_rows}. Must be 1, 2, 3, or 4'}), 400

        original_tw2_path = session.get('original_tw2_path')
        if not original_tw2_path:
            return jsonify({
                'success': False,
                'error': 'Original TW2 file path is required before saving HW Rows.'
            }), 400

        if not os.path.exists(original_tw2_path):
            return jsonify({
                'success': False,
                'error': 'Original TW2 file path is not accessible. Please validate the path and try again.'
            }), 400

        target_file = original_tw2_path

        backup_path = target_file + '.backup_hw_rows_' + datetime.now().strftime('%Y%m%d_%H%M%S')
        shutil.copy2(target_file, backup_path)

        conn = get_mdb_connection(target_file)
        cursor = conn.cursor()

        hw_rows_columns = []
        try:
            cursor.execute("SELECT * FROM tblSchedule WHERE 1=0")
            column_lookup = {(desc[0] or '').lower(): desc[0] for desc in (cursor.description or [])}

            hwrows_calc_column = column_lookup.get('hwrowscalc') or 'HWRowsCalc'
            hw_rows_columns.append(hwrows_calc_column)

            hwrows_column = column_lookup.get('hwrows')
            if hwrows_column and hwrows_column not in hw_rows_columns:
                hw_rows_columns.append(hwrows_column)

            hwrow_column = column_lookup.get('hwrow')
            if hwrow_column and hwrow_column not in hw_rows_columns:
                hw_rows_columns.append(hwrow_column)
        except Exception as e:
            print(f"HW ROWS: Unable to inspect columns: {e}")
            hw_rows_columns = hw_rows_columns or ['HWRowsCalc']

        updated_count = 0
        errors = []

        for edit in edits:
            unit_tag = edit.get('unit_tag')
            hw_rows = edit.get('hw_rows')

            if unit_tag in (None, ''):
                errors.append('Missing unit tag in edit payload')
                continue

            try:
                clean_tag = unit_tag.split('  ')[0] if '  ' in unit_tag else unit_tag

                hw_rows_value = int(hw_rows)

                set_clauses = [f'[{column}] = ?' for column in hw_rows_columns]
                params = [hw_rows_value] * len(hw_rows_columns)

                update_query = f"UPDATE tblSchedule SET {', '.join(set_clauses)} WHERE [Tag] = ?"
                params.append(clean_tag)
                cursor.execute(update_query, params)

                if cursor.rowcount > 0:
                    updated_count += 1
                    print(f"Updated HW Rows for {clean_tag}: {hw_rows_value}")
                else:
                    errors.append(f"No record found for tag: {clean_tag}")
                    print(f"No record found for tag: {clean_tag}")

            except Exception as e:
                error_msg = f"Error updating {unit_tag}: {str(e)}"
                errors.append(error_msg)
                print(error_msg)

        conn.commit()
        conn.close()

        result = {
            'success': True,
            'updated_count': updated_count,
            'backup_file': os.path.basename(backup_path),
            'target_file': os.path.basename(target_file),
            'target_path': target_file
        }

        if errors:
            result['warnings'] = errors

        return jsonify(result)

    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Failed to save HW Rows: {str(e)}'
        }), 500


@app.route('/get_updated_tw2_data', methods=['GET'])
def get_updated_tw2_data():
    """Get updated TW2 data for display"""
    try:
        if not session.get('updated_tw2_data'):
            return jsonify({'error': 'No updated TW2 data loaded'}), 400
        
        # Return the same structure as the original TW2 data viewer
        return Response(
            json.dumps({
                'success': True,
                'data': session['updated_tw2_data'],
                'columns': session.get('updated_tw2_columns', []),
                'filename': session.get('updated_tw2_filename', 'Unknown'),
                'records': session.get('updated_tw2_records', 0)
            }, cls=CustomJSONEncoder, ensure_ascii=True),
            mimetype='application/json'
        )
        
    except Exception as e:
        print(f"Error in get_updated_tw2_data: {str(e)}")
        return jsonify({'error': f'Error retrieving updated TW2 data: {str(e)}'}), 500

@app.route('/compare_performance', methods=['POST'])
def compare_performance():
    """Execute performance comparison between Excel and updated TW2 data"""
    try:
        data = request.json or {}
        mbh_lat_lower_margin = float(data.get('mbh_lat_lower_margin', 15))
        mbh_lat_upper_margin = float(data.get('mbh_lat_upper_margin', 25))
        wpd_threshold = float(data.get('wpd_threshold', 5))
        apd_threshold = float(data.get('apd_threshold', 0.25))

        # Check if required data is available
        if not session.get('excel_data'):
            return jsonify({'success': False, 'error': 'Excel data not loaded'}), 400

        reload_info = reload_tw2_data_from_disk()
        if not reload_info.get('success'):
            status_code = 404 if reload_info.get('code') == 404 else 500
            return jsonify({'success': False, 'error': 'Unable to reload TW2 data: {}'.format(reload_info.get('error'))}), status_code

        updated_tw2_data = session.get('updated_tw2_data')
        if not updated_tw2_data:
            return jsonify({'success': False, 'error': 'Updated TW2 data not loaded'}), 500

        # Perform comparison
        result = compare_performance_data(
            session['excel_data'],
            updated_tw2_data,
            mbh_lat_lower_margin=mbh_lat_lower_margin,
            mbh_lat_upper_margin=mbh_lat_upper_margin,
            wpd_threshold=wpd_threshold,
            apd_threshold=apd_threshold
        )

        if result['success']:
            return jsonify({
                'success': True,
                'data': {
                    'results': result['results'],
                    'summary': result['summary'],
                    'tw2_path': reload_info.get('path'),
                    'tw2_source': reload_info.get('source'),
                    'tw2_records': reload_info.get('row_count'),
                    'tw2_column_count': reload_info.get('column_count')
                }
            })
        else:
            return jsonify({'success': False, 'error': result['error']}), 500

    except Exception as e:
        logger.exception(f"Error in compare_performance: {str(e)}")
        return jsonify({'success': False, 'error': f'Error during comparison: {str(e)}'}), 500

@app.route('/debug_session', methods=['GET'])
def debug_session():
    """Debug endpoint to inspect current session data"""
    try:
        # Force session creation if it doesn't exist
        if not session:
            session['debug'] = 'session_created'
        
        session_keys = list(session.keys())
        debug_info = {
            'session_id': request.cookies.get('session', 'No session cookie'),
            'session_keys': session_keys,
            'session_data': {},
            'flask_session_working': True,
            'storage_type': 'filesystem'
        }
        
        # Show session data with file info
        for key in session_keys:
            if key.endswith('_path'):
                # For file paths, check if they exist
                file_path = session[key]
                debug_info['session_data'][key] = {
                    'path': file_path,
                    'exists': os.path.exists(file_path) if file_path else False,
                    'absolute_path': os.path.abspath(file_path) if file_path else None
                }
            elif key.endswith('_data'):
                # For data, just show count
                data = session[key]
                debug_info['session_data'][key] = f"[{len(data)} records]" if isinstance(data, list) else str(type(data))
            else:
                # For other data, show as is
                debug_info['session_data'][key] = session[key]
        
        return jsonify(debug_info)
        
    except Exception as e:
        return jsonify({'error': f'Debug error: {str(e)}'}), 500

@app.route('/clear_session', methods=['POST'])
def clear_session():
    """Debug endpoint to clear session data"""
    try:
        session.clear()
        return jsonify({'success': True, 'message': 'Session cleared'})
    except Exception as e:
        return jsonify({'error': f'Clear session error: {str(e)}'}), 500

@app.route('/test_large_session', methods=['POST'])
def test_large_session():
    """Test storing large data in session (Flask-Session validation)"""
    try:
        # Create a large dataset similar to TW2 data
        large_data = []
        for i in range(100):  # 100 records
            record = {}
            for j in range(50):  # 50 columns per record
                record[f'column_{j}'] = f'value_{i}_{j}_' + 'x' * 20  # Long values
            large_data.append(record)
        
        # Store in session
        session['test_large_data'] = large_data
        session['test_metadata'] = {
            'records': len(large_data),
            'columns': len(large_data[0]) if large_data else 0,
            'test_timestamp': str(datetime.now())
        }
        
        return jsonify({
            'success': True,
            'message': 'Large data stored in session successfully',
            'data_size': len(str(large_data)),
            'records': len(large_data)
        })
    except Exception as e:
        return jsonify({'error': f'Large session test error: {str(e)}'}), 500

@app.route('/validate_tw2_path', methods=['POST'])
def validate_tw2_path():
    """Validate that a TW2 file path exists and is accessible"""
    try:
        data = request.json
        file_path = _sanitize_path(data.get('path', '').strip())
        
        if not file_path:
            return jsonify({'valid': False, 'error': 'No path provided'})
        
        print(f"PATH VALIDATION: Checking path: {file_path}")
        
        # Check if path exists
        if not os.path.exists(file_path):
            return jsonify({
                'valid': False, 
                'error': f'Path not found: {file_path}',
                'details': 'File does not exist at the specified location'
            })
        
        # Check if it's a file (not directory)
        if not os.path.isfile(file_path):
            return jsonify({
                'valid': False, 
                'error': f'Path is not a file: {file_path}',
                'details': 'The specified path points to a directory, not a file'
            })
        
        # Check file extension
        if not file_path.lower().endswith(('.tw2', '.mdb')):
            return jsonify({
                'valid': False, 
                'error': 'Invalid file type',
                'details': 'File must be a .tw2 or .mdb file'
            })
        
        # Try to read the file to ensure it's accessible
        try:
            result = read_tw2_data_safe(file_path)
            if result['success']:
                return jsonify({
                    'valid': True, 
                    'message': 'Path is valid and file is readable',
                    'records': result['row_count'],
                    'columns': len(result['columns'])
                })
            else:
                return jsonify({
                    'valid': False, 
                    'error': 'File is not readable',
                    'details': f'Error reading TW2 file: {result["error"]}'
                })
        except Exception as e:
            return jsonify({
                'valid': False, 
                'error': 'File access error',
                'details': f'Unable to access file: {str(e)}'
            })
            
    except Exception as e:
        print(f"Error in path validation: {str(e)}")
        return jsonify({'valid': False, 'error': f'Validation error: {str(e)}'}), 500

@app.route('/refresh_and_compare', methods=['POST'])
def refresh_and_compare():
    """Refresh TW2 data from stored file path and automatically run comparison"""
    try:
        logger.info("REFRESH: ===== STARTING REFRESH AND COMPARE =====")
        logger.info(f"REFRESH: Session keys available: {list(session.keys())}")
        logger.info(f"REFRESH: Session ID from request: {request.cookies.get('session', 'NO SESSION COOKIE')}")

        data = request.json or {}
        mbh_lat_lower_margin = float(data.get('mbh_lat_lower_margin', 15))
        mbh_lat_upper_margin = float(data.get('mbh_lat_upper_margin', 25))
        wpd_threshold = float(data.get('wpd_threshold', 5))
        apd_threshold = float(data.get('apd_threshold', 0.25))

        request_original_path = _sanitize_path(data.get('original_path')) if data.get('original_path') else None
        preferred_paths = []

        if request_original_path:
            if os.path.exists(request_original_path):
                session['original_tw2_path'] = request_original_path
                preferred_paths.append(('original', request_original_path))
                logger.info(f"REFRESH: [OK] Using request original path: {request_original_path}")
            else:
                logger.warning(f"REFRESH: [WARNING] Request original path not accessible: {request_original_path}")

        reload_info = reload_tw2_data_from_disk(preferred_paths=preferred_paths)
        if not reload_info.get('success'):
            status_code = 404 if reload_info.get('code') == 404 else 500
            logger.error(f"REFRESH: Unable to reload TW2 data: {reload_info.get('error')}")
            return jsonify({'success': False, 'error': reload_info.get('error')}), status_code

        path_source = reload_info.get('source')
        tw2_path = reload_info.get('path')
        logger.info(f"REFRESH: Reloaded TW2 data from {tw2_path} (source: {path_source})")

        if not session.get('excel_data'):
            return jsonify({
                'success': True,
                'data': {
                    'message': 'TW2 data refreshed successfully, but Excel data not loaded for comparison',
                    'tw2_refreshed': True,
                    'comparison_available': False,
                    'path_source': path_source,
                    'tw2_path': tw2_path,
                    'tw2_records': reload_info.get('row_count'),
                    'tw2_column_count': reload_info.get('column_count'),
                    'skipped_read': False
                }
            })

        comparison_result = compare_performance_data(
            session['excel_data'],
            session['updated_tw2_data'],
            mbh_lat_lower_margin=mbh_lat_lower_margin,
            mbh_lat_upper_margin=mbh_lat_upper_margin,
            wpd_threshold=wpd_threshold,
            apd_threshold=apd_threshold
        )

        if comparison_result['success']:
            return jsonify({
                'success': True,
                'data': {
                    'message': 'TW2 data refreshed and comparison completed successfully',
                    'tw2_refreshed': True,
                    'comparison_available': True,
                    'results': comparison_result['results'],
                    'summary': comparison_result['summary'],
                    'path_source': path_source,
                    'tw2_path': tw2_path,
                    'tw2_records': reload_info.get('row_count'),
                    'tw2_column_count': reload_info.get('column_count'),
                    'skipped_read': False
                }
            })
        else:
            return jsonify({
                'success': False,
                'error': 'TW2 data refreshed but comparison failed: {}'.format(comparison_result['error']),
                'data': {
                    'tw2_refreshed': True,
                    'path_source': path_source,
                    'tw2_path': tw2_path
                }
            }), 500

    except Exception as e:
        logger.exception(f"Error in refresh_and_compare: {str(e)}")
        return jsonify({'success': False, 'error': f'Error during refresh and compare: {str(e)}'}), 500


def generate_schedule_data_excel(tw2_data, project_name):
    """Generate Schedule Data Excel report from TW2 data using template"""
    try:
        from openpyxl import load_workbook
        from copy import copy
        from openpyxl.styles import Font
        from io import BytesIO

        # Load template file
        template_path = os.path.join(os.path.dirname(__file__), 'templates', 'Schedule_Data_Template.xlsx')
        wb = load_workbook(template_path)
        ws = wb.active

        # IMPORTANT: Unmerge all cells that will be affected by row insertions and notes population
        # This must be done BEFORE we insert any rows
        merged_to_unmerge = []
        for merged_range in list(ws.merged_cells.ranges):
            # Unmerge anything in rows 6 and below (will be shifted by row insertions)
            # and anything in rows 14 and below (will be used for notes)
            if merged_range.min_row >= 6 or (merged_range.min_row >= 1 and merged_range.max_row >= 6):
                merged_to_unmerge.append(str(merged_range))

        for merged_range_str in merged_to_unmerge:
            try:
                ws.unmerge_cells(merged_range_str)
            except:
                pass

        # Update project name in row 2
        ws['A2'] = project_name

        # Get template row formatting (row 5)
        template_row = 5

        # Helper function to safely set cell value
        def safe_set_cell(cell_ref, value):
            ws[cell_ref].value = value

        # Insert data rows starting at row 5
        for i, record in enumerate(tw2_data):
            row_num = 5 + i

            # If not the first data row, insert a new row
            if i > 0:
                ws.insert_rows(row_num)

                # Copy formatting from template row to new row
                for col_num in range(1, 31):  # Columns A-AD
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(col_num)

                    source_cell = ws.cell(row=template_row, column=col_num)
                    target_cell = ws.cell(row=row_num, column=col_num)

                    # Copy formatting
                    if source_cell.font:
                        target_cell.font = copy(source_cell.font)
                    if source_cell.border:
                        target_cell.border = copy(source_cell.border)
                    if source_cell.alignment:
                        target_cell.alignment = copy(source_cell.alignment)
                    if source_cell.fill:
                        target_cell.fill = copy(source_cell.fill)
                    if source_cell.number_format:
                        target_cell.number_format = copy(source_cell.number_format)

            # Populate data columns
            try:
                safe_set_cell(f'A{row_num}', record.get('Tag', ''))
                safe_set_cell(f'F{row_num}', 'DESV')  # Model - always DESV
                safe_set_cell(f'G{row_num}', record.get('UnitSize', ''))
                safe_set_cell(f'H{row_num}', record.get('OutletSize', ''))
                safe_set_cell(f'I{row_num}', record.get('CFMDesign', ''))
                safe_set_cell(f'J{row_num}', record.get('CFMMinPrime', ''))
                safe_set_cell(f'K{row_num}', record.get('SPInlet', ''))
                safe_set_cell(f'L{row_num}', record.get('SPDownstream', ''))
                safe_set_cell(f'M{row_num}', record.get('SPMin', ''))
                safe_set_cell(f'N{row_num}', record.get('RadNCRoom', ''))
                safe_set_cell(f'O{row_num}', record.get('DisNCRoom', ''))
                safe_set_cell(f'P{row_num}', record.get('HWCFM', ''))

                if record.get('HWMBHCalc'):
                    safe_set_cell(f'Q{row_num}', round(float(record.get('HWMBHCalc', 0))))

                safe_set_cell(f'R{row_num}', record.get('HWEATCalc', ''))
                safe_set_cell(f'U{row_num}', record.get('HWEWT', ''))

                if record.get('HWLATCalc'):
                    safe_set_cell(f'V{row_num}', round(float(record.get('HWLATCalc', 0)), 1))

                if record.get('HWAPDCalc'):
                    safe_set_cell(f'W{row_num}', round(float(record.get('HWAPDCalc', 0)), 2))

                safe_set_cell(f'X{row_num}', record.get('HWGPMCalc', ''))

                if record.get('HWLWTCalc'):
                    safe_set_cell(f'Y{row_num}', round(float(record.get('HWLWTCalc', 0)), 1))

                if record.get('HWPDCalc'):
                    safe_set_cell(f'Z{row_num}', round(float(record.get('HWPDCalc', 0)), 2))

                hw_rows = record.get('HWRowsCalc') or record.get('HWRows', '')
                control_hand = record.get('ControlHand', '')
                if hw_rows:
                    safe_set_cell(f'AA{row_num}', f"{hw_rows}-{control_hand}")

                safe_set_cell(f'AB{row_num}', record.get('HWFPI', ''))
                safe_set_cell(f'AC{row_num}', record.get('ControlHand', ''))

                # Re-merge cells for this data row to match original template structure
                ws.merge_cells(f'A{row_num}:B{row_num}')
                ws.merge_cells(f'C{row_num}:E{row_num}')
                ws.merge_cells(f'R{row_num}:S{row_num}')

            except Exception as e:
                logger.error(f"Error processing row for tag {record.get('Tag', 'Unknown')}: {str(e)}")
                continue

        # Place notes section after data
        notes_start_row = 5 + len(tw2_data) + 2

        # Get fluid type info from first record
        fluid_type = tw2_data[0].get('FluidType', '') if tw2_data else ''
        pct_glycol = tw2_data[0].get('PctGlycol', 40) if tw2_data else 40

        if fluid_type == 'EG':
            fluid_description = f"{pct_glycol}% Ethylene Glycol"
        elif fluid_type == 'PG':
            fluid_description = f"{pct_glycol}% Propylene Glycol"
        elif fluid_type == 'Water' or fluid_type == '':
            fluid_description = "100% Water"
        else:
            fluid_description = f"{pct_glycol}% {fluid_type}"

        # Notes content
        notes = [
            ("Notes:", "1. Selections are based on Titus as Manufacturer."),
            (None, "2. All performance based on tests conducted in accordance with ASHRAE 130-2008 and AHRI 880-2011."),
            (None, "3. All NC levels determined using AHRI 885-2008 Appendix E."),
            (None, "4. All airflow, pressure losses and heating performance values have been corrected for altitude."),
            (None, "5. Units of measure: dimensions (in), airflow (cfm), water flow (gpm), air pressure (in wg), water head losses (ft) and temperatures (degF)."),
            (None, "6. Water pressure drop (WPd) units is in ft. water."),
            (None, f"7. Hot water performance based on {fluid_description}.")
        ]

        current_row = notes_start_row
        for label, note in notes:
            # Set row height to match other note rows
            ws.row_dimensions[current_row].height = 15

            if label:
                safe_set_cell(f'B{current_row}', label)
                ws[f'B{current_row}'].font = Font(name='Arial', size=8, bold=True)
                safe_set_cell(f'E{current_row}', note)
                # Merge notes label cells
                try:
                    ws.merge_cells(f'B{current_row}:C{current_row}')
                except:
                    pass
            else:
                safe_set_cell(f'E{current_row}', note)

            # Apply 8pt Arial font to note content
            ws[f'E{current_row}'].font = Font(name='Arial', size=8)

            # Merge notes content cells through column V for more room
            try:
                ws.merge_cells(f'E{current_row}:V{current_row}')
            except:
                pass
            current_row += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    except Exception as e:
        logger.exception(f"Error generating schedule data Excel: {str(e)}")
        raise


@app.route('/export_schedule_data', methods=['POST'])
def export_schedule_data():
    """Export TW2 data as Schedule Data Excel report"""
    try:
        updated_tw2_data = session.get('updated_tw2_data') or session.get('tw2_data')
        if not updated_tw2_data:
            return jsonify({'success': False, 'error': 'TW2 data not loaded'}), 400

        tw2_path = session.get('original_tw2_path') or session.get('tw2_file_path', '')
        project_name = 'VAV Schedule Data'

        if tw2_path:
            import os
            filename = os.path.splitext(os.path.basename(tw2_path))[0]
            if ' - ' in filename:
                project_name = filename.split(' - ')[0].strip()
            else:
                project_name = filename

        excel_file = generate_schedule_data_excel(updated_tw2_data, project_name)

        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Schedule_Data_{timestamp}.xlsx"

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logger.exception(f"Error in export_schedule_data: {str(e)}")
        return jsonify({'success': False, 'error': f'Error generating report: {str(e)}'}), 500


if __name__ == '__main__':
    app.run(debug=True, port=5004)
